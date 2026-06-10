"""
Download Address XLS for Spain (ESP) rank-by-rank, then merge into one Excel.

Why by rank: the full Spain dataset is too large for a single request.
Splitting by rank keeps each download small and lets the run be resumed.

Steps
-----
1. For every rank in rank.txt (268 ranks):
   - Search /search with country_code=ESP, rank_id=<id>, show_active unchecked
   - If the "Address xls" button appears  → download → save to
     downloads/spain/<SANITIZED_RANK>.xls
   - If the button is absent (no results)  → skip
2. Merge all per-rank XLS files into one Excel:
   - Parse every file in downloads/spain/
   - Deduplicate on seaman_id (keep first occurrence)
   - Write to downloads/spain_combined.xlsx

Flags
-----
    --download-only    only run step 1 (skip the merge)
    --merge-only       only run step 2 (skip the download)
    (default: run both steps)

.env must contain: CREW_BASE_URL, CREW_USERNAME, CREW_PASSWORD
"""

import asyncio
import io
import os
import re
import sys

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from html.parser import HTMLParser
from pathlib import Path
from typing import Optional

import xlrd
from dotenv import load_dotenv
from playwright.async_api import async_playwright, Page

load_dotenv()

BASE_URL    = os.getenv("CREW_BASE_URL", "").rstrip("/")
CI_USERNAME = os.getenv("CREW_USERNAME", "")
CI_PASSWORD = os.getenv("CREW_PASSWORD", "")

SPAIN_CODE   = "ESP"
SPAIN_DIR    = Path("downloads/spain")
COMBINED_OUT = Path(f"downloads/{SPAIN_CODE}.xlsx")
REQUEST_DELAY = 2.0
BATCH_SIZE    = 50     # rows per Supabase insert (if you later push to DB)

DO_DOWNLOAD = "--merge-only"  not in sys.argv
DO_MERGE    = "--download-only" not in sys.argv

_BTN = 'input[type="button"][value="Address xls"]'


# ── rank-map builder (reused from scraper.py) ───────────────────────────────────

class _RankParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self._in  = False
        self._val: Optional[str] = None
        self.ranks: dict[str, int] = {}

    def handle_starttag(self, tag, attrs):
        d = dict(attrs)
        if tag == "select" and d.get("id") == "rank_id":
            self._in = True
        if self._in and tag == "option":
            self._val = d.get("value", "0")

    def handle_endtag(self, tag):
        if tag == "select":
            self._in = False

    def handle_data(self, data):
        if self._in and self._val is not None:
            name = data.strip()
            if name and self._val not in ("0", ""):
                self.ranks[name] = int(self._val)
            self._val = None


def build_rank_map(html_path: str) -> dict[str, int]:
    with open(html_path, encoding="utf-8") as f:
        html = f.read()
    p = _RankParser()
    p.feed(html)
    return p.ranks


def safe_filename(rank_name: str) -> str:
    """Convert a rank name to a safe filename stem."""
    return re.sub(r'[\\/:*?"<>|]+', "_", rank_name).strip("_ ")


# ── download step ───────────────────────────────────────────────────────────────

async def download_rank(page: Page, rank_name: str, rank_id: int, dest: Path) -> str:
    """
    Search for Spain + rank, then download the Address XLS if the button appears.

    Returns: "ok" | "no_button" | "failed"
    """
    await page.goto(f"{BASE_URL}/search", wait_until="domcontentloaded", timeout=30_000)
    await asyncio.sleep(0.5)

    # Set rank
    await page.select_option("select#rank_id", value=str(rank_id))

    # Set country = ESP
    await page.fill("input#country_code", SPAIN_CODE)
    await page.keyboard.press("Escape")
    await asyncio.sleep(0.3)

    # Uncheck "Only active" to get ALL seafarers
    cb = page.locator("input#show_active")
    if await cb.is_checked():
        await cb.click()
    await asyncio.sleep(0.2)

    # Run the search so the results table renders
    try:
        async with page.expect_navigation(wait_until="networkidle", timeout=30_000):
            await page.locator('input[type="submit"][value="Search"]').first.click()
    except Exception:
        try:
            await page.wait_for_load_state("networkidle", timeout=15_000)
        except Exception:
            pass

    # Only download if the button is present (i.e. there are results)
    btn = page.locator(_BTN)
    if not await btn.is_visible():
        return "no_button"

    try:
        async with page.expect_download(timeout=60_000) as dl_info:
            await btn.click()
        download = await dl_info.value
        await download.save_as(str(dest))
        return "ok"
    except Exception as exc:
        print(f"\n    WARNING download failed [{rank_name}]: {exc}", file=sys.stderr)
        return "failed"


async def run_download(rank_map: dict[str, int]) -> None:
    SPAIN_DIR.mkdir(parents=True, exist_ok=True)

    rank_list = list(rank_map.items())
    print(f"Ranks to process : {len(rank_list)}")
    print(f"Saving to        : {SPAIN_DIR.resolve()}\n")

    ok_count     = 0
    skip_count   = 0
    no_btn_count = 0
    fail_count   = 0

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=False, slow_mo=80)
        ctx     = await browser.new_context(
            accept_downloads=True,
            viewport={"width": 1280, "height": 800},
        )
        page = await ctx.new_page()

        # ── Login ──────────────────────────────────────────────────────────────
        print("Logging in ...")
        await page.goto(f"{BASE_URL}/login", wait_until="domcontentloaded")
        await page.fill('input[name="login"]',    CI_USERNAME)
        await page.fill('input[name="password"]', CI_PASSWORD)
        async with page.expect_navigation(wait_until="domcontentloaded", timeout=20_000):
            await page.click('input[name="submit_bttn"]')
        await page.wait_for_load_state("networkidle", timeout=15_000)

        try:
            await page.wait_for_selector('a[href="/logout"]', timeout=8_000)
            print("  Logged in OK.\n")
        except Exception:
            await page.screenshot(path="login_debug.png")
            raise RuntimeError("Login failed – logout link not found.")

        # ── Per-rank download loop ─────────────────────────────────────────────
        for i, (rname, rid) in enumerate(rank_list, 1):
            fname = safe_filename(rname)
            dest  = SPAIN_DIR / f"{fname}.xls"

            label = f"  [{i:3}/{len(rank_list)}] {rname:<30}"

            if dest.exists():
                print(f"{label} SKIP  (already downloaded)")
                skip_count += 1
                continue

            print(f"{label} ...", end=" ", flush=True)
            result = await download_rank(page, rname, rid, dest)

            if result == "ok":
                size_kb = dest.stat().st_size // 1024
                print(f"OK  ({size_kb} KB)")
                ok_count += 1
            elif result == "no_button":
                print("no results – skipped")
                no_btn_count += 1
            else:
                print("FAILED")
                fail_count += 1

            await asyncio.sleep(REQUEST_DELAY)

        # ── Logout ─────────────────────────────────────────────────────────────
        print("\nLogging out ...")
        await page.goto(f"{BASE_URL}/logout", wait_until="domcontentloaded")
        await asyncio.sleep(1.0)
        await browser.close()

    print(f"\n=== Download complete ===")
    print(f"  Downloaded : {ok_count}")
    print(f"  Skipped    : {skip_count}  (file already existed)")
    print(f"  No results : {no_btn_count}  (button not shown)")
    print(f"  Failed     : {fail_count}")


# ── merge step ──────────────────────────────────────────────────────────────────

def _cell_str(ws, row: int, col: int) -> str:
    v = ws.cell_value(row, col)
    if v is None or v == "":
        return ""
    if isinstance(v, float):
        i = int(v)
        return str(i) if i == v else str(v)
    return str(v).strip()


def parse_xls(path: Path) -> list[dict]:
    """Parse one Address XLS (rows start at index 3) and return list of dicts."""
    rank_name = path.stem  # filename = sanitized rank name
    wb = xlrd.open_workbook(
        str(path),
        ignore_workbook_corruption=True,
        logfile=io.StringIO(),
    )
    ws = wb.sheet_by_index(0)
    rows: list[dict] = []
    for r in range(3, ws.nrows):
        raw_id = ws.cell_value(r, 0)
        if not raw_id:
            continue
        rows.append({
            "seaman_id":    int(float(raw_id)),
            "rank":         _cell_str(ws, r, 1)  or None,
            "name":         _cell_str(ws, r, 2)  or None,
            "surname":      _cell_str(ws, r, 3)  or None,
            "relation":     _cell_str(ws, r, 4)  or None,
            "country":      _cell_str(ws, r, 5)  or None,
            "country_code": SPAIN_CODE,
            "city":         _cell_str(ws, r, 6)  or None,
            "county":       _cell_str(ws, r, 7)  or None,
            "street":       _cell_str(ws, r, 8)  or None,
            "postal_code":  _cell_str(ws, r, 9)  or None,
            "email":        _cell_str(ws, r, 10) or None,
            "phone":        _cell_str(ws, r, 11) or None,
            "mobile":       _cell_str(ws, r, 12) or None,
            "payroll_id":   _cell_str(ws, r, 13) or None,
        })
    return rows


_ILLEGAL_CHARS = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")


def _clean(value):
    """Strip control characters that openpyxl cannot write to Excel cells."""
    if isinstance(value, str):
        return _ILLEGAL_CHARS.sub("", value)
    return value


def run_merge() -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    xls_files = sorted(SPAIN_DIR.glob("*.xls"))
    if not xls_files:
        print(f"No .xls files found in {SPAIN_DIR} – nothing to merge.")
        return

    print(f"\nMerging {len(xls_files)} rank files ...")

    all_rows: list[dict] = []
    seen_ids: set[int]   = set()
    dupes = 0

    for path in xls_files:
        rows = parse_xls(path)
        before = len(all_rows)
        for row in rows:
            sid = row["seaman_id"]
            if sid in seen_ids:
                dupes += 1
                continue
            seen_ids.add(sid)
            all_rows.append(row)
        added = len(all_rows) - before
        print(f"  {path.name:<35} {len(rows):>5,} rows  (+{added:,} unique)")

    print(f"\n  Total unique seafarers : {len(all_rows):,}")
    print(f"  Duplicate rows removed : {dupes:,}")

    # ── Write Excel ────────────────────────────────────────────────────────────
    COLUMNS = [
        "seaman_id", "rank", "name", "surname", "relation",
        "country", "country_code", "city", "county", "street",
        "postal_code", "email", "phone", "mobile", "payroll_id",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Spain Seafarers"

    # Header row styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align

    # Data rows
    for row_idx, row in enumerate(all_rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=_clean(row.get(col_name)))

    # Auto-fit column widths (estimate)
    col_widths = {col: len(col) for col in COLUMNS}
    for row in all_rows[:500]:  # sample first 500 rows for width estimate
        for col_name, val in row.items():
            if val and col_name in col_widths:
                col_widths[col_name] = max(col_widths[col_name], min(len(str(val)), 50))
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = col_widths[col_name] + 2

    # Freeze header row
    ws.freeze_panes = "A2"

    COMBINED_OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(COMBINED_OUT))
    print(f"\n  Saved: {COMBINED_OUT.resolve()}")
    print(f"  Rows : {len(all_rows):,}  Columns: {len(COLUMNS)}")


# ── main ────────────────────────────────────────────────────────────────────────

async def main() -> None:
    if DO_DOWNLOAD:
        missing = [k for k, v in {
            "CREW_BASE_URL": BASE_URL,
            "CREW_USERNAME": CI_USERNAME,
            "CREW_PASSWORD": CI_PASSWORD,
        }.items() if not v]
        if missing:
            sys.exit(f"Missing env vars: {', '.join(missing)}")

    rank_map = build_rank_map("search.html")
    # Filter to only ranks that exist in rank.txt
    with open("rank.txt", encoding="utf-8") as f:
        rank_names = {l.strip() for l in f if l.strip()}
    rank_map = {k: v for k, v in rank_map.items() if k in rank_names}
    print(f"=== Spain Address XLS — download by rank ===")
    print(f"Ranks available : {len(rank_map)}\n")

    if DO_DOWNLOAD:
        await run_download(rank_map)

    if DO_MERGE:
        run_merge()


if __name__ == "__main__":
    asyncio.run(main())
