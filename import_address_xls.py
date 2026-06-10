"""
Import Address XLS files from downloads/ into the Supabase `seafarer_addresses` table.

Each XLS file is named <ISO_CODE>.xls (e.g. PHL.xls).  The file layout is:
    Row 0 : "Address" title (ignored)
    Row 1 : empty (ignored)
    Row 2 : column headers — ID, Rank, Name, Surname, Relation, Country,
                             City, County, Street, Postal code, E-mail,
                             Phone, Mobile, Payroll ID
    Row 3+: data

Run schema_address.sql in the Supabase SQL Editor before the first import.

Usage:
    python import_address_xls.py              # import all files in downloads/
    python import_address_xls.py --test       # first 3 files only
    python import_address_xls.py PHL IDN MMR  # specific ISO codes

.env must contain: SUPABASE_URL, SUPABASE_SERVICE_KEY
"""

import io
import os
import sys
import glob

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from datetime import datetime, timezone
from pathlib import Path

import xlrd
from dotenv import load_dotenv
from supabase import create_client, Client

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")
DOWNLOAD_DIR = Path("downloads")
BATCH_SIZE   = 200   # rows per Supabase upsert call

TEST_MODE    = "--test" in sys.argv

# Optional positional ISO codes:  python import_address_xls.py PHL IDN
FILTER_CODES = {a.upper() for a in sys.argv[1:] if not a.startswith("-")}


# ── XLS parser ──────────────────────────────────────────────────────────────────

def _cell_str(ws, row: int, col: int) -> str:
    """Return cell value as a stripped string (empty string for blanks)."""
    v = ws.cell_value(row, col)
    if v is None or v == "":
        return ""
    if isinstance(v, float):
        # xlrd returns integers as float (e.g. 1004.0 → "1004")
        i = int(v)
        return str(i) if i == v else str(v)
    return str(v).strip()


def parse_xls(path: Path) -> list[dict]:
    """
    Parse a raw Address .xls file (layout: title row 0, blank row 1,
    header row 2, data from row 3). Returns a list of row dicts.
    """
    country_code = path.stem.upper()

    wb = xlrd.open_workbook(
        str(path),
        ignore_workbook_corruption=True,
        logfile=io.StringIO(),
    )
    ws = wb.sheet_by_index(0)

    rows: list[dict] = []
    for r in range(3, ws.nrows):
        raw_id = ws.cell_value(r, 0)
        if not raw_id:
            continue
        rows.append({
            "seaman_id":    int(float(raw_id)),
            "rank":         _cell_str(ws, r, 1)  or None,
            "name":         _cell_str(ws, r, 2)  or None,
            "surname":      _cell_str(ws, r, 3)  or None,
            "relation":     _cell_str(ws, r, 4)  or None,
            "country":      _cell_str(ws, r, 5)  or None,
            "country_code": country_code,
            "city":         _cell_str(ws, r, 6)  or None,
            "county":       _cell_str(ws, r, 7)  or None,
            "street":       _cell_str(ws, r, 8)  or None,
            "postal_code":  _cell_str(ws, r, 9)  or None,
            "email":        _cell_str(ws, r, 10) or None,
            "phone":        _cell_str(ws, r, 11) or None,
            "mobile":       _cell_str(ws, r, 12) or None,
            "payroll_id":   _cell_str(ws, r, 13) or None,
        })

    return rows


def parse_xlsx(path: Path) -> list[dict]:
    """
    Parse a combined .xlsx file produced by download_spain_by_rank.py
    (layout: header row 1, data from row 2). Returns a list of row dicts.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    # Map header names to column indices (1-based)
    header = {cell.value: cell.column for cell in next(ws.iter_rows(min_row=1, max_row=1))}

    def get(row, col_name):
        idx = header.get(col_name)
        if idx is None:
            return None
        v = row[idx - 1].value
        if v is None:
            return None
        return str(v).strip() or None

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2):
        raw_id = row[header["seaman_id"] - 1].value
        if not raw_id:
            continue
        rows.append({
            "seaman_id":    int(float(raw_id)),
            "rank":         get(row, "rank"),
            "name":         get(row, "name"),
            "surname":      get(row, "surname"),
            "relation":     get(row, "relation"),
            "country":      get(row, "country"),
            "country_code": get(row, "country_code"),
            "city":         get(row, "city"),
            "county":       get(row, "county"),
            "street":       get(row, "street"),
            "postal_code":  get(row, "postal_code"),
            "email":        get(row, "email"),
            "phone":        get(row, "phone"),
            "mobile":       get(row, "mobile"),
            "payroll_id":   get(row, "payroll_id"),
        })

    wb.close()
    return rows


# ── Supabase upsert ─────────────────────────────────────────────────────────────

def upsert_rows(sb: Client, rows: list[dict]) -> None:
    """Upsert rows in batches; on conflict (seaman_id) update all columns."""
    for start in range(0, len(rows), BATCH_SIZE):
        batch = rows[start : start + BATCH_SIZE]
        sb.table("seafarer_addresses").upsert(
            batch,
            on_conflict="seaman_id",
        ).execute()


# ── main ────────────────────────────────────────────────────────────────────────

def main() -> None:
    missing = [k for k, v in {
        "SUPABASE_URL":         SUPABASE_URL,
        "SUPABASE_SERVICE_KEY": SUPABASE_KEY,
    }.items() if not v]
    if missing:
        sys.exit(f"Missing env vars: {', '.join(missing)}")

    # Collect files to process.
    # If both <CODE>.xls and <CODE>.xlsx exist for the same country, prefer
    # the .xlsx (the deduplicated combined file, e.g. ESP.xlsx from
    # download_spain_by_rank.py) and skip the raw .xls.
    xls_files  = {f.stem.upper(): f for f in DOWNLOAD_DIR.glob("*.xls")}
    xlsx_files = {f.stem.upper(): f for f in DOWNLOAD_DIR.glob("*.xlsx")}
    merged = {**xls_files, **xlsx_files}   # xlsx wins on collision
    all_files = sorted(merged.values(), key=lambda f: f.stem.upper())

    if not all_files:
        sys.exit(f"No .xls / .xlsx files found in {DOWNLOAD_DIR.resolve()}")

    if FILTER_CODES:
        files = [f for f in all_files if f.stem.upper() in FILTER_CODES]
        missing_codes = FILTER_CODES - {f.stem.upper() for f in files}
        if missing_codes:
            print(f"WARNING: no file found for: {', '.join(sorted(missing_codes))}")
    elif TEST_MODE:
        files = all_files[:3]
        print(f"*** TEST MODE – first 3 files ***")
    else:
        files = all_files

    print(f"Files to import : {len(files)}")
    print(f"Supabase table  : seafarer_addresses\n")

    sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

    total_rows   = 0
    ok_files     = 0
    fail_files   = 0
    started_at   = datetime.now(timezone.utc)

    for i, path in enumerate(files, 1):
        code = path.stem.upper()
        print(f"  [{i:3}/{len(files)}] {code} ...", end=" ", flush=True)

        try:
            rows = parse_xlsx(path) if path.suffix == ".xlsx" else parse_xls(path)
            if not rows:
                print("0 rows – skipped")
                continue

            upsert_rows(sb, rows)
            print(f"{len(rows):>6,} rows upserted")
            total_rows += len(rows)
            ok_files   += 1

        except Exception as exc:
            print(f"FAILED – {exc}", file=sys.stderr)
            fail_files += 1

    elapsed = (datetime.now(timezone.utc) - started_at).total_seconds()
    print(f"\n=== Done in {elapsed:.1f}s ===")
    print(f"  Files OK    : {ok_files}")
    print(f"  Files failed: {fail_files}")
    print(f"  Rows upserted: {total_rows:,}")


if __name__ == "__main__":
    main()
